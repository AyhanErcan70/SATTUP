import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from app.styles import COLORS, PATHS, COMPANY_NAME
import re

def create_excel(table_widget, report_title="Genel Rapor", username="Admin", parent=None, meta=None):
    """
    QTableWidget verilerini Excel'e dönüştürür.
    TEK SAYFAYA SIĞDIRMA GARANTİLİ VERSİYON.
    """
    
    # 1. Kayıt Yeri Seçimi
    meta_d = meta if isinstance(meta, dict) else {}
    is_bulk_puantaj = "Toplu Puantaj" in str(report_title or "")
    cust_for_name = ""
    try:
        if is_bulk_puantaj:
            cust_for_name = str(meta_d.get("customer_name") or "").strip()
    except Exception:
        cust_for_name = ""
    if cust_for_name:
        default_name = f"{cust_for_name}_{datetime.now().strftime('%d%m%Y')}.xlsx"
    else:
        default_name = f"{report_title}_{datetime.now().strftime('%d%m%Y')}.xlsx"
    
    file_name, _ = QFileDialog.getSaveFileName(
        parent,
        "Excel Olarak Kaydet",
        os.path.join(os.path.expanduser("~/Desktop"), default_name),
        "Excel Dosyası (*.xlsx)"
    )
    
    if not file_name:
        return

    try:
        src_col_count = table_widget.columnCount()
        is_bulk_puantaj = "Toplu Puantaj" in str(report_title or "")

        template_path = None
        if is_bulk_puantaj:
            try:
                repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
                cand = os.path.join(repo_root, "ui", "Toplu_puantaj_sablon.xlsx")
                if os.path.exists(cand):
                    template_path = cand
            except Exception:
                template_path = None

        if template_path:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            # openpyxl does not preserve existing embedded images when loading/saving.
            # Re-add the company logo so the template's header area stays visually consistent.
            try:
                if os.path.exists(PATHS["logo"]):
                    img = XLImage(PATHS["logo"])
                    img.width = 120
                    img.height = 50
                    ws.add_image(img, "A1") # type: ignore
            except Exception:
                pass

            def _hex2(x: int) -> str:
                try:
                    return f"{int(x) & 0xFF:02X}"
                except Exception:
                    return "00"

            def _qcolor_to_rgb_hex(c) -> str | None:
                try:
                    if c is None:
                        return None
                    # QColor
                    if hasattr(c, "red") and hasattr(c, "green") and hasattr(c, "blue"):
                        return f"{_hex2(c.red())}{_hex2(c.green())}{_hex2(c.blue())}"
                    # QBrush
                    if hasattr(c, "color"):
                        qc = c.color()
                        return f"{_hex2(qc.red())}{_hex2(qc.green())}{_hex2(qc.blue())}"
                except Exception:
                    return None
                return None

            def _qt_item_bg_hex(item) -> str | None:
                if item is None:
                    return None
                try:
                    bg = item.data(Qt.ItemDataRole.BackgroundRole)
                    hx = _qcolor_to_rgb_hex(bg)
                    if hx:
                        return hx
                except Exception:
                    pass
                try:
                    hx = _qcolor_to_rgb_hex(item.background())
                    if hx:
                        return hx
                except Exception:
                    pass
                return None

            def _px_to_points(px: int | float) -> float:
                # Heuristic conversion (Qt px -> Excel points). Tuned for typical Windows DPI.
                try:
                    return max(0.0, float(px) * 0.75)
                except Exception:
                    return 0.0

            def _tr_float(s0: str) -> float | None:
                try:
                    t = str(s0 or "").strip()
                    if not t:
                        return None
                    t = t.replace("₺", "").replace("TL", "")
                    t = t.replace(" ", "")
                    # Remove thousands separator and convert decimal comma
                    t = t.replace(".", "").replace(",", ".")
                    return float(t)
                except Exception:
                    return None

            def _tr_int(s0: str) -> int | None:
                try:
                    t = str(s0 or "").strip()
                    if not t:
                        return None
                    if re.fullmatch(r"[-+]?\d+", t):
                        return int(t)
                except Exception:
                    return None
                return None

            def _anchor_cell(row: int, col: int):
                try:
                    for mr in ws.merged_cells.ranges:
                        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                            return ws.cell(row=mr.min_row, column=mr.min_col) # type: ignore
                except Exception:
                    pass
                return ws.cell(row=row, column=col) # type: ignore

            def _merge_is_safe(r1: int, c1: int, r2: int, c2: int) -> bool:
                # Avoid creating overlapping/invalid merges on top of the template's existing merges.
                try:
                    nr1, nc1 = int(min(r1, r2)), int(min(c1, c2))
                    nr2, nc2 = int(max(r1, r2)), int(max(c1, c2))
                except Exception:
                    return False

                try:
                    existing = list(ws.merged_cells.ranges)
                except Exception:
                    existing = []

                for mr in existing:
                    try:
                        er1, ec1 = int(mr.min_row), int(mr.min_col)
                        er2, ec2 = int(mr.max_row), int(mr.max_col)
                    except Exception:
                        continue

                    # Identical merge already exists.
                    if er1 == nr1 and ec1 == nc1 and er2 == nr2 and ec2 == nc2:
                        return False

                    # Overlap check.
                    rows_overlap = not (nr2 < er1 or nr1 > er2)
                    cols_overlap = not (nc2 < ec1 or nc1 > ec2)
                    if rows_overlap and cols_overlap:
                        return False

                return True

            def _qt_span_anchor(r: int, c: int) -> tuple[int, int]:
                # If (r,c) is inside a span, find the top-left anchor of that span.
                try:
                    for rr in range(int(r), -1, -1):
                        try:
                            rs = int(table_widget.rowSpan(int(rr), int(c)) or 1)
                        except Exception:
                            rs = 1
                        if rs <= 1:
                            continue
                        if int(rr) + int(rs) <= int(r):
                            continue
                        for cc in range(int(c), -1, -1):
                            try:
                                cs = int(table_widget.columnSpan(int(rr), int(cc)) or 1)
                            except Exception:
                                cs = 1
                            if cs <= 1:
                                continue
                            if int(cc) + int(cs) <= int(c):
                                continue
                            return int(rr), int(cc)
                except Exception:
                    pass
                return int(r), int(c)

            def _norm_hdr(x: str) -> str:
                return re.sub(r"\s+", " ", str(x or "").strip()).casefold()

            max_col = int(ws.max_column or 0)
            header_rows = (5, 6) if is_bulk_puantaj else None
            if not header_rows:
                max_row = int(ws.max_row or 0)
                header_row = None
                for r in range(1, max_row + 1):
                    row_vals = []
                    for c in range(1, max_col + 1):
                        v = ws.cell(row=r, column=c).value
                        row_vals.append(_norm_hdr(v) if v is not None else "")
                    hit = 0
                    for key in ["güzergah", "guzergah", "şoför", "sofor", "hareket"]:
                        if any((key in s) for s in row_vals):
                            hit += 1
                    if hit >= 2:
                        header_row = r
                        break
                if header_row is None:
                    raise RuntimeError("Şablonda başlık satırı bulunamadı")
                header_rows = (int(header_row),)

            tmpl_headers = {}
            for c in range(1, int(ws.max_column or 0) + 1):
                parts = []
                for rr in header_rows or ():
                    try:
                        v = _anchor_cell(int(rr), int(c)).value
                        nv = _norm_hdr(v)
                        if nv:
                            parts.append(nv)
                    except Exception:
                        pass
                joined = _norm_hdr(" ".join([p for p in parts if p]))
                if joined and joined not in tmpl_headers:
                    tmpl_headers[joined] = int(c)

            tmpl_export_cols = []
            for c in range(1, int(ws.max_column or 0) + 1):
                has_hdr = False
                for rr in header_rows or ():
                    try:
                        v = _anchor_cell(int(rr), int(c)).value
                        if _norm_hdr(v):
                            has_hdr = True
                            break
                    except Exception:
                        pass
                if has_hdr:
                    tmpl_export_cols.append(int(c))

            src_headers = [table_widget.horizontalHeaderItem(i).text() for i in range(src_col_count)]

            export_src_cols = list(range(src_col_count))
            if is_bulk_puantaj:
                try:
                    def _norm_hdr2(x: str) -> str:
                        return str(x or "").replace("\n", " ").strip().casefold()

                    drop_idx = None
                    for i, h in enumerate(src_headers):
                        if _norm_hdr2(h) == "duraklar":
                            drop_idx = i
                            break
                    if drop_idx is not None:
                        export_src_cols = [i for i in export_src_cols if i != int(drop_idx)]
                except Exception:
                    pass

            src_to_tmpl_col = {}

            if is_bulk_puantaj:
                try:
                    def _tr_month_name(m: int) -> str:
                        mm = int(m)
                        names = {
                            1: "OCAK",
                            2: "ŞUBAT",
                            3: "MART",
                            4: "NİSAN",
                            5: "MAYIS",
                            6: "HAZİRAN",
                            7: "TEMMUZ",
                            8: "AĞUSTOS",
                            9: "EYLÜL",
                            10: "EKİM",
                            11: "KASIM",
                            12: "ARALIK",
                        }
                        return names.get(mm, str(mm))

                    meta_d2 = meta if isinstance(meta, dict) else {}
                    cust = str(meta_d2.get("customer_name") or "").strip()
                    if cust:
                        _anchor_cell(3, 4).value = cust

                    ym = str(meta_d2.get("month") or "").strip()
                    m = re.match(r"^\s*(20\d{2})[-./](\d{1,2})\s*$", ym)
                    if m:
                        _anchor_cell(2, 40).value = str(m.group(1))
                        _anchor_cell(3, 40).value = _tr_month_name(int(m.group(2)))

                    tmpl_fixed = list(tmpl_export_cols[:6])
                    tmpl_days = list(tmpl_export_cols[6:37])
                    tmpl_sum = list(tmpl_export_cols[37:40])

                    day_src_cols: list[tuple[int, int]] = []
                    sum_src_cols: list[int] = []
                    fixed_src_cols: list[int] = []

                    for src_c in export_src_cols:
                        try:
                            h = str(src_headers[int(src_c)] or "").strip()
                        except Exception:
                            h = ""
                        mday = re.match(r"^\s*(\d{1,2})\b", h)
                        if mday:
                            try:
                                dn = int(mday.group(1))
                            except Exception:
                                dn = -1
                            if 1 <= dn <= 31:
                                day_src_cols.append((int(dn), int(src_c)))
                                continue

                        hn = _norm_hdr(h)
                        if any(
                            k in hn
                            for k in [
                                "toplam sefer",
                                "sefer say",
                                "sefer sayisi",
                                "sefer sayısı",
                                "sefer baş",
                                "sefer bas",
                                "toplam tutar",
                                "tutar",
                                "fiyat",
                            ]
                        ):
                            sum_src_cols.append(int(src_c))
                            continue

                        fixed_src_cols.append(int(src_c))

                    day_src_cols.sort(key=lambda x: x[0])
                    src_days = [c for _dn, c in day_src_cols]

                    if len(sum_src_cols) < 3 and len(export_src_cols) >= 3:
                        sum_src_cols = list(export_src_cols[-3:])
                    else:
                        sum_src_cols = [c for c in export_src_cols if c in set(sum_src_cols)]
                    src_sum = list(sum_src_cols[-3:])

                    src_fixed = list(fixed_src_cols[:6])

                    if (
                        len(tmpl_fixed) == 6
                        and len(tmpl_days) == 31
                        and len(tmpl_sum) == 3
                        and len(src_fixed) == 6
                        and len(src_sum) == 3
                    ):
                        for i in range(6):
                            src_to_tmpl_col[int(src_fixed[i])] = int(tmpl_fixed[i])

                        for i in range(min(len(src_days), 31)):
                            src_to_tmpl_col[int(src_days[i])] = int(tmpl_days[i])

                        for i in range(3):
                            src_to_tmpl_col[int(src_sum[i])] = int(tmpl_sum[i])
                except Exception:
                    src_to_tmpl_col = {}

            start_data_row = 7 if is_bulk_puantaj else (int((header_rows or (0,))[0]) + 1)

            # Remove any pre-existing merges in the template's data body so our Qt-derived merges
            # can be applied cleanly.
            try:
                if is_bulk_puantaj and (src_to_tmpl_col or {}):
                    body_r1 = int(start_data_row)
                    body_r2 = int(start_data_row) + int(table_widget.rowCount()) - 1
                    body_c1 = int(min((src_to_tmpl_col or {}).values() or [1]))
                    body_c2 = int(max((src_to_tmpl_col or {}).values() or [body_c1]))
                    to_unmerge = []
                    for mr in list(ws.merged_cells.ranges):
                        try:
                            if int(mr.max_row) < body_r1 or int(mr.min_row) > body_r2:
                                continue
                            if int(mr.max_col) < body_c1 or int(mr.min_col) > body_c2:
                                continue
                            to_unmerge.append(str(mr))
                        except Exception:
                            continue
                    for rng in to_unmerge:
                        try:
                            ws.unmerge_cells(rng) # type: ignore
                        except Exception:
                            pass
            except Exception:
                pass

            row_count = table_widget.rowCount()
            for r in range(int(row_count)):
                out_r = int(start_data_row) + int(r)

                # Preserve row heights from the Qt table.
                try:
                    ws.row_dimensions[int(out_r)].height = _px_to_points(int(table_widget.rowHeight(int(r)))) # type: ignore
                except Exception:
                    pass
                for src_c, tmpl_c in (src_to_tmpl_col or {}).items():
                    ar, ac = _qt_span_anchor(int(r), int(src_c))
                    item = table_widget.item(int(ar), int(ac))
                    val = item.text() if item else ""
                    if is_bulk_puantaj and int(tmpl_c) == int(tmpl_headers.get(_norm_hdr("GÜZERGAH")) or -1):
                        val = str(val or "")
                    xl_cell = _anchor_cell(int(out_r), int(tmpl_c))
                    # Write numeric cells as numbers to avoid Excel "text" warnings.
                    wrote_number = False
                    try:
                        if is_bulk_puantaj:
                            # Day qty columns in template are 7..37 (1-based, 31 days)
                            if 7 <= int(tmpl_c) <= 37:
                                ival = _tr_int(val)
                                if ival is not None:
                                    xl_cell.value = int(ival)
                                    xl_cell.number_format = "0"
                                    wrote_number = True
                            # Totals area: 38..40 (TOPLAM SEFER SAYISI, SEFER BAŞI FİYAT, TOPLAM TUTAR)
                            if (not wrote_number) and int(tmpl_c) in (38, 39, 40):
                                if int(tmpl_c) == 38:
                                    ival = _tr_int(val)
                                    if ival is not None:
                                        xl_cell.value = int(ival)
                                        xl_cell.number_format = "0"
                                        wrote_number = True
                                else:
                                    fval = _tr_float(val)
                                    if fval is not None:
                                        xl_cell.value = float(fval)
                                        xl_cell.number_format = "#,##0.00"
                                        wrote_number = True
                    except Exception:
                        wrote_number = False

                    if not wrote_number:
                        xl_cell.value = val

                    # Basic alignment: keep template defaults, but ensure merged/text-heavy cells wrap.
                    try:
                        if is_bulk_puantaj and int(tmpl_c) in (1, 2, 3, 4, 5, 6, 38, 39, 40):
                            xl_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    except Exception:
                        pass

                    # Background fills: only weekend/holiday/override (qty coloring explicitly not required).
                    try:
                        bg_hex = _qt_item_bg_hex(item)
                        # Only allow known semantic backgrounds to avoid accidental black fills.
                        if bg_hex and str(bg_hex).strip().upper() in {"CFCFCF", "F8C291"}:
                            xl_cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
                    except Exception:
                        pass

                    # For overrides/notes: add an Excel comment using tooltip text (if any).
                    try:
                        tip = str(item.toolTip() if item is not None else "").strip()
                        if not tip:
                            # If the cell is visually marked as override but tooltip isn't populated,
                            # still leave a minimal note.
                            try:
                                bg_hex = _qt_item_bg_hex(item)
                                if str(bg_hex or "").strip().upper() == "F8C291":
                                    tip = "Override"
                            except Exception:
                                tip = ""
                        if tip:
                            xl_cell.comment = Comment(tip, "SATTUP")
                    except Exception:
                        pass

            # Apply QTableWidget spans as Excel merged cells (data body only).
            try:
                if is_bulk_puantaj and (src_to_tmpl_col or {}):
                    for r in range(int(row_count)):
                        out_r = int(start_data_row) + int(r)
                        for src_c, tmpl_c0 in (src_to_tmpl_col or {}).items():
                            rs = int(table_widget.rowSpan(int(r), int(src_c)) or 1)
                            cs = int(table_widget.columnSpan(int(r), int(src_c)) or 1)
                            if rs <= 1 and cs <= 1:
                                continue
                            # Merge only from the anchor cell.
                            ar, ac = _qt_span_anchor(int(r), int(src_c))
                            if int(ar) != int(r) or int(ac) != int(src_c):
                                continue

                            try:
                                tmpl_cols = []
                                for cc in range(int(src_c), int(src_c) + int(cs)):
                                    if int(cc) in (src_to_tmpl_col or {}):
                                        tmpl_cols.append(int((src_to_tmpl_col or {})[int(cc)]))
                                if not tmpl_cols:
                                    continue
                                min_c = int(min(tmpl_cols))
                                max_c = int(max(tmpl_cols))
                                r1 = int(out_r)
                                c1 = int(min_c)
                                r2 = int(out_r) + int(rs) - 1
                                c2 = int(max_c)
                                if _merge_is_safe(int(r1), int(c1), int(r2), int(c2)):
                                    ws.merge_cells(
                                        start_row=int(r1),
                                        start_column=int(c1),
                                        end_row=int(r2),
                                        end_column=int(c2),
                                    ) # type: ignore
                            except Exception:
                                pass
            except Exception:
                pass

            wb.save(file_name)
            QMessageBox.information(parent, "Başarılı", f"Excel dosyası oluşturuldu:\n{file_name}")
            os.startfile(file_name)
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapor" # type: ignore

        # --- 2. Sayfa Düzeni ve SIĞDIRMA AYARLARI ---
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE # type: ignore
        ws.page_setup.paperSize = ws.PAPERSIZE_A4 # type: ignore
        
        # --- KRİTİK DÜZELTME BAŞLANGICI ---
        # Excel'e "Sayfa yapısı ayarlarını kullan" emrini veriyoruz
        ws.sheet_properties.pageSetUpPr.fitToPage = True  # type: ignore
        
        # Genişliği KESİN OLARAK 1 sayfaya sığdır
        ws.page_setup.fitToWidth = 1 # type: ignore
        ws.page_setup.fitToHeight = False # type: ignore # Yükseklik serbest (otomatik artsın)
        # --- KRİTİK DÜZELTME BİTİŞİ ---

        ws.print_options.horizontalCentered = True # type: ignore
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.3, footer=0.3) # type: ignore

        # --- 3. Logo ---
        if os.path.exists(PATHS["logo"]):
            try:
                img = XLImage(PATHS["logo"])
                img.width = 120
                img.height = 50
                ws.add_image(img, "A1") # type: ignore
            except Exception:
                pass

        headers = [table_widget.horizontalHeaderItem(i).text() for i in range(src_col_count)]
        is_bulk_puantaj = "Toplu Puantaj" in str(report_title or "")

        export_src_cols = list(range(src_col_count))
        if is_bulk_puantaj:
            try:
                def _norm_hdr(x: str) -> str:
                    return str(x or "").replace("\n", " ").strip().casefold()

                drop_idx = None
                for i, h in enumerate(headers):
                    if _norm_hdr(h) == "duraklar":
                        drop_idx = i
                        break
                if drop_idx is not None:
                    export_src_cols = [i for i in export_src_cols if i != int(drop_idx)]
            except Exception:
                pass

        col_count = len(export_src_cols)
        last_col_letter = get_column_letter(col_count)

        # --- 4. Başlıklar ---
        title_cell = ws["A2"] # type: ignore
        title_cell.value = f"{COMPANY_NAME} - {report_title}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells(f"A2:{last_col_letter}2") # type: ignore

        # --- 5. Tablo Başlıkları ---
        headers = [headers[i] for i in export_src_cols]

        if is_bulk_puantaj:
            def _simplify_day_header(h: str) -> str:
                txt = str(h or "").strip()
                if not txt:
                    return txt
                # Common formats: "1\nPz", "1 Pz", "01\nPt".
                m = re.match(r"^\s*(\d{1,2})\b", txt)
                if not m:
                    return txt
                try:
                    day = int(m.group(1))
                except Exception:
                    return txt
                if 1 <= day <= 31:
                    return str(day)
                return txt

            headers = [_simplify_day_header(h) for h in headers]
        
        header_fill = PatternFill(start_color=COLORS["header_fill"], end_color=COLORS["header_fill"], fill_type="solid")
        header_font = Font(bold=True, color=COLORS["header_text"])
        thin_border = Border(left=Side(style="thin", color=COLORS["border"]), 
                             right=Side(style="thin", color=COLORS["border"]), 
                             top=Side(style="thin", color=COLORS["border"]), 
                             bottom=Side(style="thin", color=COLORS["border"]))

        start_row = 5
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name) # type: ignore
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            if "Toplu Puantaj" in str(report_title or ""):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        try:
            if "Toplu Puantaj" in str(report_title or ""):
                ws.row_dimensions[start_row].height = 32 # type: ignore
        except Exception:
            pass

        # --- 6. Veriler ---
        row_count = table_widget.rowCount()
        for r in range(row_count):
            for out_c, src_c in enumerate(export_src_cols):
                item = table_widget.item(r, int(src_c))
                val = item.text() if item else ""
                
                cell = ws.cell(row=start_row + 1 + r, column=out_c + 1, value=val) # type: ignore
                cell.border = thin_border
                h_align = "left" if out_c in [3, 6] else "center"
                if is_bulk_puantaj and (out_c + 1) in [1, 2, 3, 4]:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal=h_align, vertical="center")

        # --- 7. Sütun Genişlikleri (Hafif Revize Edildi) ---
        # 1:Kod, 2:Tür, 3:TCKN, 4:AdSoyad, 5:Görev, 6:GSM, 7:Email, 8:Kan, 9:Durum
        custom_widths = {
            1: 13,  # Kod
            2: 20,  # Tür 
            3: 15,  # TCKN
            4: 30,  # Ad Soyad
            5: 20,  # Görev
            6: 18,  # GSM
            7: 35,  # Email
            8: 10,  # Kan
            9: 10   # Durum
        }

        is_bulk_puantaj = "Toplu Puantaj" in str(report_title or "")
        bulk_widths = {
            1: 3.5546875,
            2: 10.44140625,
            3: 9.33203125,
            4: 10.21875,
            5: 7.21875,
            6: 9.44140625,
            7: 2.88671875,
        }

        for col_idx, col_name in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            if is_bulk_puantaj:
                try:
                    if col_idx in bulk_widths:
                        ws.column_dimensions[col_letter].width = float(bulk_widths[col_idx]) # type: ignore
                        continue
                except Exception:
                    pass

                try:
                    if col_idx >= 7 and col_idx <= (col_count - 3):
                        ws.column_dimensions[col_letter].width = 2.88671875 # type: ignore
                        continue
                except Exception:
                    pass

                try:
                    last4_start = max(1, int(col_count) - 3)
                    if col_idx == last4_start:
                        ws.column_dimensions[col_letter].width = 7.33203125 # type: ignore
                    elif col_idx == last4_start + 1:
                        ws.column_dimensions[col_letter].width = 9.5546875 # type: ignore
                    elif col_idx == last4_start + 2:
                        ws.column_dimensions[col_letter].width = 10.33203125 # type: ignore
                    elif col_idx == last4_start + 3:
                        ws.column_dimensions[col_letter].width = 8.88671875 # type: ignore
                    else:
                        ws.column_dimensions[col_letter].width = 15 # type: ignore
                except Exception:
                    ws.column_dimensions[col_letter].width = 15 # type: ignore
            else:
                if col_idx in custom_widths:
                    ws.column_dimensions[col_letter].width = custom_widths[col_idx] # type: ignore
                else:
                    ws.column_dimensions[col_letter].width = 15 # type: ignore

        try:
            if is_bulk_puantaj:
                start_data_row = start_row + 1
                end_data_row = start_row + int(row_count)
                for merge_col in [1, 2, 3, 4, 5]:
                    r0 = start_data_row
                    while r0 <= end_data_row:
                        v0 = ws.cell(row=r0, column=merge_col).value # type: ignore
                        v0s = re.sub(r"\s+", " ", str(v0 or "").strip()).casefold()
                        r1 = r0
                        while r1 + 1 <= end_data_row:
                            v1 = ws.cell(row=r1 + 1, column=merge_col).value # type: ignore
                            v1s = re.sub(r"\s+", " ", str(v1 or "").strip()).casefold()
                            if v1s != v0s:
                                break
                            r1 += 1
                        if v0s and r1 > r0:
                            ws.merge_cells(start_row=r0, start_column=merge_col, end_row=r1, end_column=merge_col) # type: ignore
                            if merge_col in [1, 2, 3, 4]:
                                try:
                                    top_cell = ws.cell(row=r0, column=merge_col) # type: ignore
                                    top_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                                except Exception:
                                    pass
                        r0 = r1 + 1
        except Exception:
            pass

        # --- 8. Yazdırma Alanı ---
        ws.print_area = f"A1:{last_col_letter}{ws.max_row}" # type: ignore
        ws.print_title_rows = '5:5' # type: ignore

        wb.save(file_name)
        QMessageBox.information(parent, "Başarılı", f"Excel dosyası oluşturuldu:\n{file_name}")
        os.startfile(file_name)

    except Exception as e:
        QMessageBox.critical(parent, "Hata", f"Excel hatası:\n{e}")