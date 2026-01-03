import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from app.styles import COLORS, PATHS, COMPANY_NAME

def create_excel(table_widget, report_title="Genel Rapor", username="Admin", parent=None):
    """
    QTableWidget verilerini Excel'e dönüştürür.
    TEK SAYFAYA SIĞDIRMA GARANTİLİ VERSİYON.
    """
    
    # 1. Kayıt Yeri Seçimi
    default_name = f"{report_title}_{datetime.now().strftime('%d%m%Y')}.xlsx"
    file_name, _ = QFileDialog.getSaveFileName(
        parent,
        "Excel Olarak Kaydet",
        os.path.join(os.path.expanduser("~/Desktop"), default_name),
        "Excel Dosyası (*.xlsx)"
    )
    
    if not file_name:
        return

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapor" # type: ignore

        # --- 2. Sayfa Düzeni ve SIĞDIRMA AYARLARI ---
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE # type: ignore
        ws.page_setup.paperSize = ws.PAPERSIZE_A4 # type: ignore
        
        # --- KRİTİK DÜZELTME BAŞLANGICI ---
        # Excel'e "Sayfa yapısı ayarlarını kullan" emrini veriyoruz
        ws.sheet_properties.pageSetUpPr.fitToPage = True  # type: ignore
        
        # Genişliği KESİN OLARAK 1 sayfaya sığdır
        ws.page_setup.fitToWidth = 1 # type: ignore
        ws.page_setup.fitToHeight = False # type: ignore # Yükseklik serbest (otomatik artsın)
        # --- KRİTİK DÜZELTME BİTİŞİ ---

        ws.print_options.horizontalCentered = True # type: ignore
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.3, footer=0.3) # type: ignore

        # --- 3. Logo ---
        if os.path.exists(PATHS["logo"]):
            try:
                img = XLImage(PATHS["logo"])
                img.width = 120
                img.height = 50
                ws.add_image(img, "A1") # type: ignore
            except Exception:
                pass

        col_count = table_widget.columnCount()
        last_col_letter = get_column_letter(col_count)

        # --- 4. Başlıklar ---
        title_cell = ws["A2"] # type: ignore
        title_cell.value = f"{COMPANY_NAME} - {report_title}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells(f"A2:{last_col_letter}2") # type: ignore

        # --- 5. Tablo Başlıkları ---
        headers = [table_widget.horizontalHeaderItem(i).text() for i in range(col_count)]
        
        header_fill = PatternFill(start_color=COLORS["header_fill"], end_color=COLORS["header_fill"], fill_type="solid")
        header_font = Font(bold=True, color=COLORS["header_text"])
        thin_border = Border(left=Side(style="thin", color=COLORS["border"]), 
                             right=Side(style="thin", color=COLORS["border"]), 
                             top=Side(style="thin", color=COLORS["border"]), 
                             bottom=Side(style="thin", color=COLORS["border"]))

        start_row = 5
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name) # type: ignore
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # --- 6. Veriler ---
        row_count = table_widget.rowCount()
        for r in range(row_count):
            for c in range(col_count):
                item = table_widget.item(r, c)
                val = item.text() if item else ""
                
                cell = ws.cell(row=start_row + 1 + r, column=c + 1, value=val) # type: ignore
                cell.border = thin_border
                h_align = "left" if c in [3, 6] else "center"
                cell.alignment = Alignment(horizontal=h_align, vertical="center")

        # --- 7. Sütun Genişlikleri (Hafif Revize Edildi) ---
        # 1:Kod, 2:Tür, 3:TCKN, 4:AdSoyad, 5:Görev, 6:GSM, 7:Email, 8:Kan, 9:Durum
        custom_widths = {
            1: 13,  # Kod
            2: 20,  # Tür 
            3: 15,  # TCKN
            4: 30,  # Ad Soyad
            5: 20,  # Görev
            6: 18,  # GSM
            7: 35,  # Email
            8: 10,  # Kan
            9: 10   # Durum
        }

        for col_idx, col_name in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            if col_idx in custom_widths:
                ws.column_dimensions[col_letter].width = custom_widths[col_idx] # type: ignore
            else:
                ws.column_dimensions[col_letter].width = 15 # type: ignore

        # --- 8. Yazdırma Alanı ---
        ws.print_area = f"A1:{last_col_letter}{ws.max_row}" # type: ignore
        ws.print_title_rows = '5:5' # type: ignore

        wb.save(file_name)
        QMessageBox.information(parent, "Başarılı", f"Excel dosyası oluşturuldu:\n{file_name}")
        os.startfile(file_name)

    except Exception as e:
        QMessageBox.critical(parent, "Hata", f"Excel hatası:\n{e}")