# C:\ELBEK\app\export_utils.py
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from PyQt6.QtWidgets import QTableWidget, QFileDialog, QMessageBox, QWidget
from PyQt6.QtCore import QDate, Qt
from fpdf import FPDF # PDF için fpdf kütüphanesi
from PIL import Image # Filigran için Pillow kütüphanesi

# =====================================================
# === 1. EXCEL OLUŞTURMA MODÜLÜ (Çalışan Kodunuz) =====
# =====================================================

def export_to_excel(table_widget: QTableWidget, report_title="Rapor", username="Kullanıcı"):
    """QTableWidget verisini Excel dosyasına (xlsx) aktarır."""
    if table_widget.rowCount() == 0:
        QMessageBox.warning(None, "Uyarı", "Aktarılacak veri bulunmamaktadır!")
        return False

    filename, _ = QFileDialog.getSaveFileName(
        None, "Excel Dosyası Kaydet", "", "Excel Dosyaları (*.xlsx)"
    )
    if not filename:
        return False

    try:
        # Kodun geri kalanı aynı kalır (Sayfa düzeni, başlıklar, veriler)
        wb = Workbook()
        ws = wb.active
        ws.title = report_title
        
        # ... (Sizin Excel stil ayarlarınız ve logo kodlarınız buraya gelir) ...
        # (Kullanıcının image_56f068.png ile gösterdiği tüm kodlar buraya entegre edildi varsayılır)

        ws["C2"] = report_title
        ws["C2"].font = Font(bold=True, size=14)
        # ... (Kolon başlıkları ve verilerin yazılması) ...
        
        # --- Kolon Başlıkları (Tkinter'da olduğu gibi QTableWidget'tan çekiyoruz) ---
        headers = []
        for col in range(table_widget.columnCount()):
            header_text = table_widget.horizontalHeaderItem(col).text()
            headers.append(header_text)
            ws.cell(row=5, column=col + 1, value=header_text).font = Font(bold=True)
            
        # --- Veri Satırları ---
        for row in range(table_widget.rowCount()):
            for col in range(table_widget.columnCount()):
                item = table_widget.item(row, col)
                value = item.text() if item else ""
                ws.cell(row=row + 6, column=col + 1, value=value)
        
        # ... (Geri kalan boyutlandırma ve kaydetme mantığı) ...

        wb.save(filename)
        QMessageBox.information(None, "Başarılı", f"Veriler Excel'e aktarıldı:\n{filename}")
        return True

    except Exception as e:
        QMessageBox.critical(None, "Hata", f"Excel'e aktarırken bir hata oluştu: {str(e)}")
        return False


# =====================================================
# === 2. PDF OLUŞTURMA MODÜLÜ (Stil Entegrasyonu) =====
# =====================================================

# C:\ELBEK\app\export_utils.py dosyasına, export_to_pdf fonksiyonu yerine:

# export_utils.py dosyası, export_to_pdf fonksiyonu:

# Not: Dosyanın başında 'from fpdf2 import FPDF' olmalı.

# C:\ELBEK\app\export_utils.py dosyası, export_to_pdf fonksiyonu:

import os
import sys
from datetime import datetime
from openpyxl import Workbook
# ... (Diğer importlar) ...
from PyQt6.QtWidgets import QTableWidget, QFileDialog, QMessageBox, QWidget
from PyQt6.QtCore import QDate, Qt
from fpdf import FPDF # <-- fpdf2'yi kullandığımızdan emin ol
from PIL import Image # Filigran için Pillow (Emin olmak için pip install Pillow yapıldı)


def export_to_pdf(table_widget: QTableWidget, report_title="Personel Raporu", user_name="Kullanıcı"):
    """
    QTableWidget verilerini PDF'e dönüştürür (UTF-8/Türkçe uyumlu) ve filigran ekler.
    """
    if table_widget.rowCount() == 0:
        QMessageBox.warning(None, "Uyarı", "Aktarılacak veri bulunmamaktadır!")
        return False
        
    save_path, _ = QFileDialog.getSaveFileName(
        None, "PDF Dosyası Kaydet", f"{report_title}.pdf", "PDF Dosyaları (*.pdf)"
    )
    if not save_path:
        return False

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.alias_nb_pages()
    pdf.add_page()
    
    # --- FONT VE LOGO YOLLARI ---
    fonts_dir = r"C:\ELBEK\fonts"
    icons_dir = r"C:\ELBEK\icons"
    dejavu_regular_path = os.path.join(fonts_dir, "DejaVuSans.ttf")
    dejavu_bold_path = os.path.join(fonts_dir, "DejaVuSans-Bold.ttf") # <-- ARTIK AYRI BİR DOSYA
    dejavu_italic_path = os.path.join(fonts_dir, "DejaVuSans-Oblique.ttf")
    logo_path = os.path.join(icons_dir, "logo.png")
    faded_logo_path = os.path.join(icons_dir, "logo_faded.png")
    font_name = "Arial" 
    
    font_name = "Arial" 

    if os.path.exists(dejavu_regular_path):
        try:
            # A. Normal Stili Yükle (Zorunlu)
            pdf.add_font("DejaVu", "", dejavu_regular_path, uni=True) 
            
            # B. Bold Stili Yükle (Bağımsız IF)
            if os.path.exists(dejavu_bold_path):
                pdf.add_font("DejaVu", "B", dejavu_bold_path, uni=True) 
            
            # C. ITALIC Stili Yükle (Bağımsız IF - Hatanın Kaynağı)
            if os.path.exists(dejavu_italic_path):
                pdf.add_font("DejaVu", "I", dejavu_italic_path, uni=True) 
            else:
                # Yedek: Italic dosyası yoksa, normal fontu Italic olarak kaydet (Hata almamak için)
                pdf.add_font("DejaVu", "I", dejavu_regular_path, uni=True) 
                
            font_name = "DejaVu" # Başarılı yüklemeyi onaylar
            
        except Exception as e:
            print(f"FPDF Font Yükleme Hatası: {e}")
            pass
            
    # --- 2. BAŞLIK VE LOGO ---
    if os.path.exists(logo_path):
        pdf.image(logo_path, x=10, y=8, w=40)

    pdf.set_xy(55, 15)
    pdf.set_font(font_name, "B", 14) 
    # Sadece font adını değil, stringin kendisinin de UTF-8 uyumlu olduğundan emin olmalıyız.
    pdf.cell(0, 10, f"Sakarya Asil Tur Taşımacılık Hizmetleri — {report_title}", align="C", ln=1)
    pdf.ln(8)
    
    # --- 3. Sütun Başlıkları ---
    headers = []
    for col in range(table_widget.columnCount()):
        headers.append(table_widget.horizontalHeaderItem(col).text())
        
    col_widths = [20, 30, 30, 70, 20, 30, 45, 20, 15] 

    # Başlık Stili
    pdf.set_font(font_name, "B", 9)
    pdf.set_fill_color(60, 60, 60)
    pdf.set_text_color(255, 255, 255)
    
    for i, h in enumerate(headers):
        width = col_widths[i] if i < len(col_widths) else 30
        pdf.cell(width, 8, str(h), border=1, align="C", fill=True)
    pdf.ln()

    # --- 4. Veri Satırları ---
    pdf.set_font(font_name, "", 8) 
    pdf.set_text_color(0, 0, 0)
    fill = False
    
    for row in range(table_widget.rowCount()):
        data_row = []
        for col in range(table_widget.columnCount()):
            item = table_widget.item(row, col)
            data_row.append(item.text() if item else "")

        pdf.set_fill_color(245, 245, 245) if fill else pdf.set_fill_color(255, 255, 255)
        
        for i, text in enumerate(data_row):
            width = col_widths[i] if i < len(col_widths) else 30
            align = "L" if i in (3, 6) else "C" 
            
            pdf.cell(width, 7, str(text), border=1, align=align, fill=True) 
            
        pdf.ln()
        fill = not fill

    # --- 5. FİLİGRAN (WATERMARK) --
    if os.path.exists(logo_path):
        # Filigran görüntüsü oluşturma (Eski kodundan)
        if not os.path.exists(faded_logo_path):
            with Image.open(logo_path).convert("RGBA") as im:
                alpha = im.split()[3]
                alpha = alpha.point(lambda p: p * 0.18)
                im.putalpha(alpha)
                im.save(faded_logo_path)
        
        # Sayfanın ortasına yerleştirme
        x_center = pdf.w / 2
        y_center = pdf.h / 2
        pdf.image(faded_logo_path, x=x_center - 75, y=y_center - 35, w=150)

    # --- 6. FOOTER ---
    
    pdf.set_auto_page_break(auto=False)
    tarih = datetime.now().strftime("%d.%m.%Y %H:%M")
    footer_text = f"{user_name} — {tarih} — Sayfa {pdf.page_no()}/{{nb}}"
    
    page_height = pdf.h - 15
    pdf.set_y(page_height - 10)
    pdf.set_draw_color(0, 0, 0)
    pdf.set_line_width(0.4)
    pdf.line(10, page_height - 10, pdf.w - 10, page_height - 10) # Çizgi çek
    
    pdf.set_font(family="DejaVu",style="I",size=7)
    pdf.cell(0, 8, "Sakarya Asil Tur Taşımacılık Hizmetleri", align="L")
    pdf.cell(-10, 8, footer_text, align="R") 


    # --- 7. Kaydetme ---
    try:
        pdf.output(save_path)
        QMessageBox.information(None, "Başarılı", f"PDF başarıyla kaydedildi:\n{save_path}")
        return True
    except Exception as e:
        QMessageBox.critical(None, "Hata", f"PDF aktarımı başarısız oldu: {str(e)}")
        return False