from pathlib import Path

import openpyxl


def _cell_str(v):
    if v is None:
        return ""
    s = str(v)
    s = s.replace("\n", " ").strip()
    return s


def _is_header_like(v: object) -> bool:
    if v in (None, ""):
        return False
    if isinstance(v, str):
        t = v.strip()
        if not t:
            return False
        # avoid rows that are mostly formulas
        if t.startswith("="):
            return False
        return True
    return False


def _col_letter(idx_1: int) -> str:
    # 1-indexed
    n = int(idx_1)
    s = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def _detect_header_row(ws, maxc: int) -> tuple[int, int]:
    # Prefer a row with many header-like (string, not formula) cells.
    best_r = 1
    best_score = -1
    for r in range(1, 81):
        nonempty = 0
        header_like = 0
        for c in range(1, maxc + 1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                nonempty += 1
            if _is_header_like(v):
                header_like += 1
        score = header_like * 3 + nonempty
        if score > best_score:
            best_score = score
            best_r = r
    return best_r, best_score


def main():
    path = Path(r"c:\Users\ayhan\SATTUP\ui\hakedis_ornek.xlsx")
    if not path.exists():
        raise SystemExit(f"File not found: {path}")

    wb = openpyxl.load_workbook(str(path), data_only=False)
    wb_val = openpyxl.load_workbook(str(path), data_only=True)
    print("SHEETS:")
    for s in wb.sheetnames:
        print(f"- {s}")

    for name in wb.sheetnames:
        ws = wb[name]
        ws_val = wb_val[name]
        maxc = min(int(ws.max_column or 1), 120)
        print("\n===", name, "===")

        header_row, header_score = _detect_header_row(ws, min(maxc, 80))
        # collect headers
        headers: list[tuple[str, str]] = []
        for c in range(1, min(maxc, 80) + 1):
            h = _cell_str(ws.cell(header_row, c).value)
            if h:
                headers.append((_col_letter(c), h))

        print(f"HEADER_ROW={header_row} score={header_score} headers={len(headers)}")
        if headers:
            print("HDR:")
            print(" | ".join([f"{col}:{txt}" for col, txt in headers[:40]]))

        # Show first 8 non-empty rows after header row (values workbook preferred)
        print("SAMPLE_ROWS:")
        sample_shown = 0
        for r in range(header_row + 1, header_row + 250):
            row_vals = [ws_val.cell(r, c).value for c in range(1, min(maxc, 40) + 1)]
            if not any(v not in (None, "") for v in row_vals):
                continue
            out = []
            for c, v in enumerate(row_vals, start=1):
                if v in (None, ""):
                    continue
                out.append(f"{_col_letter(c)}={_cell_str(v)[:60]}")
            if out:
                print(f"{r}: " + " | ".join(out[:30]))
                sample_shown += 1
            if sample_shown >= 8:
                break


if __name__ == "__main__":
    main()
